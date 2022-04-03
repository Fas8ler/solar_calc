import matplotlib.pyplot as plt
import numpy as np
from scipy.interpolate import make_interp_spline, BSpline
from openpyxl import load_workbook
from scipy.optimize import minimize
import string
import sys
import os


# funkcija izravcuna razliko med zacetno in koncno napolnjenostjo baterije za dane podatke dataT
#parIn vsebuje 2 parmetra, povrsina celic Sc in zacetno napolnjenost baterije E0
# vhodni podatki so se kapaciteta baterij Cb, podatki o obsevanju in porabi dataT, st. ur v dataT dimT, pltQ doloci, ali se rise graf ali ne.
def dEfunc(parIn,Cb,dataT,dimT,pltQ):
    
    Sc=parIn[0]
    E0=parIn[1]
    datB1 = np.zeros([dimT+1])
    datB1[-1] = E0

    #spreminjanje napolnjenosti baterije izracunamo z zanko. Zgornja meja je dolocena z Cb
    for j in range(dimT+1):
        datB1[j] = np.minimum(datB1[np.mod(j-1,dimT+1)] + Sc*dataT[np.mod(j,dimT),4] - dataT[np.mod(j,dimT),5],Cb)

    #iz izracunane napolnjenosti dolocimo razliko med zacetnim in koncnim stanjem dE in minimalno napolnjenost v bateriji Emin
    dE = datB1[-1] - datB1[0]
    Emin = np.min(datB1)
    if pltQ:
        plt.plot(datB1)
        plt.show()
    #print("[Sc,E0]=")
    print(parIn)
    print("vals=")
    print(dE)
    print(Emin)
    #print("-------------------")
    
    # funkcija vrne vsoto absolutnih vrednosti dE in Emin. Ce je sistem usklanjen, mora biti to enako 0.
    return np.abs(dE)+np.abs(Emin)


power_produce = load_workbook(filename="Data_Year_Kem.xlsx")
sheet_prod = power_produce.active
power_consume = load_workbook(filename="DV_Data_1.xlsx")
sheet_con = power_consume.active


mypath = os.path.dirname(os.path.realpath(__file__))


print(sheet_prod  )
i0=12
i1=8795
dimT = i1-i0

dataT = np.zeros((dimT,6))

print(dimT)
#print(bla)


#v tej zanki uvozimo podatke
for i in range(dimT):
	v = sheet_prod.cell(row=i+12, column=1).value
	v1 = sheet_prod.cell(row=i+12, column=3).value
	v2 = sheet_con.cell(row=i+2, column=6).value
	vNek = sheet_con.cell(row=i+2, column=7).value
	vHidro = sheet_con.cell(row=i+2, column=9).value
	#print(vHidro)
	#print(i)
	s1 = v.split(":")
	#print(s1)	
	date=int(s1[0])-20160000
	day = np.mod(date,100)
	month = (date-day)/100
	#print(date)
	#print(day)
	#print(month)
	#dataT[i]=[i,month,day,round(0.01*int(s1[1])),float(v1),float(v2)-float(vNek)-float(vHidro)]
	dataT[i]=[i,month,day,round(0.01*int(s1[1])),float(v1),float(v2)]
	#print(dataT[i])
	



# zacenta povrsina celic
SSC = np.sum(dataT[:,5])/np.sum(dataT[:,4])



# za orientacijsko vrednost kolicine baterij naredimo standardni racun za napolnjenost baterije...
datB = np.zeros([dimT+1])
datB[-1] = 0.0
for j in range(dimT+1):
    datB[j] = datB[np.mod(j-1,dimT)] + SSC*dataT[np.mod(j,dimT),4] - dataT[np.mod(j,dimT),5]


#... in dolocimo kapaciteto baterij
Cbat = np.max(datB) - np.min(datB)
dE = datB[-1] - datB[0]

#zacetni predlog za zacetno napolnjenost baterij, kapaciteto baterij in povrsino celic
Ebat0 = - np.min(datB)
Cbat0 = Cbat
SSC0 = SSC



# kolicino baterij omejimo na fiksno vrednost
Cbat1 = 0.05*Cbat0

# predlog za povrsino in napolnjenost pravimo v spremenljivko par0
#par0 = [100*SSC0,0.9*Ebat0]
par0 = [5*SSC0,0.9*Cbat1]

print("Bat. kapaciteta = "+str(Cbat1)+" MWh")
#minimizacijski algoritem prilagaja vrednosti SC in Ebat0 (v par0) toliko casa, da najde vrednosti, pri katerih je zacetna in koncna
#napolnjnost baterije enaka in da napolnjenost nikoli ne pade pod 0.
MM = minimize(dEfunc,par0,args=(Cbat1,dataT,dimT,False), method='Nelder-Mead')

#optimizirane vrednosti preberemo iz MM
parS = MM.x
print(MM)
print("Opt. povrsina celic = "+str(parS[0])+" km^2")
print("Zacetna napolnjenost baterij = "+str(parS[1])+" MWh")

cenaY = np.round(parS[0] *10**6 *80 *10**-9,3) + np.round(Cbat1 *10**3 *140*10**-9,3)
print("Cena = "+str(np.round(cenaY,3))+" milijard eur")

#narisemo lahko se graf

dEfunc(parS,Cbat1,dataT,dimT,True)

print(stop)


#tule spodaj je se poskus zanke, ki avtomatsko racuna povrino celic za razlicne kapacitete baterij.
nnMax = 101
dataOpt = np.zeros((nnMax,4))

for i in range(nnMax):

    Cbat1 = (0.1-0.001*i)*Cbat
    dataOpt[i,0] = Cbat1
    par0 = [SSC0,Ebat0]

    MM = minimize(dEfunc,par0,args=(Cbat1,dataT,dimT,False), method='Nelder-Mead')

    parS = MM.x
    dataOpt[i,1:3] = parS
    #print(MM)
    print("Bat. kapaciteta = "+str(Cbat1)+" MWh")
    print("Opt. povrsina celic = "+str(parS[0])+" km^2")
    #dEfunc(parS,Cbat1,dataT,dimT,True)

    cenaY = np.round(parS[0] *10**6 *80 *10**-9,3) + np.round(Cbat1 *10**3 *140*10**-9,3)
    dataOpt[i,3] = cenaY
    print("Cena = "+str(np.round(cenaY,3))+" milijard eur")

    parF=open(mypath+"/dataOpt/dataCalc.txt",'a')
    parF.write(str(parS[0]) + "\t")
    parF.write(str(Cbat1) + "\t")
    #parF.write(str(parS[1]) + "\t")
    parF.write(str(cenaY) + "\n")
    parF.close()


print(dataOpt)
np.savetxt(mypath+"/dataOpt/data1.txt",dataOpt,fmt='%.3e', delimiter='\t', newline='\n')
